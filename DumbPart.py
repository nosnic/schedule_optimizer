import json
import openpyxl


class TableWorker:
    class Lession:
        def __init__(self, name, information, day, time):
            self.name = name
            self.teacher = information("teacher")
            self.session = information("session")
            self.profile = information("profile")
            self.notable = information("notable")
            self.lecture = information("lecture")
            self.day = day
            self.time = time

    def readonly_table(self, TablefileName):
        wookbookTable = openpyxl.load_workbook(TablefileName)
        worksheetTable = wookbookTable.active
        rows = worksheetTable.max_row - 1
        columns = worksheetTable.max_column
        table = []
        for i in range(1, worksheetTable.max_row):
            day = []
            j = 0
            for col in worksheetTable.iter_cols(1, worksheetTable.max_column):
                if col[i].value != None:
                    day.append(col[i].value)
                else:
                    day.append("")
            table.append(day)
        return table


    def read_table(self, TablefileName, additionalInfoFileName):
        with open(additionalInfoFileName, 'r', encoding='utf-8') as f:
            data = json.load(f)
        wookbookTable = openpyxl.load_workbook(TablefileName)
        worksheetTable = wookbookTable.active
        rows = worksheetTable.max_row - 1
        columns = worksheetTable.max_column
        table = []
        for i in range(1, worksheetTable.max_row):
            day = []
            j = 0
            for col in worksheetTable.iter_cols(1, worksheetTable.max_column):
                j += 1
                if (col[i].value != None):
                    data[col[i].value]["day"] = j
                    data[col[i].value]["time"] = i
                    day.append(data[col[i].value])
            table.append(day)
        return table

    def create_info_table(self, table, additionalInfoFileName, saving_lessions=[]):
        result = []
        flag = True
        error = ""
        with open(additionalInfoFileName, 'r', encoding='utf-8') as f:
            data = json.load(f)
        valid_keys = list(data)
        for i in range(len(table)):
            day = []
            for j in range(len(table[i])):
                if (table[i][j] != ""):
                    if table[i][j] in valid_keys:
                        if [i, j] in saving_lessions:
                            data[table[i][j]] = data["Максимум"]
                        data[table[i][j]]["day"] = j
                        data[table[i][j]]["time"] = i
                        day.append(data[table[i][j]])
                    else:
                        flag = False
                        error = table[i][j]
            result.append(day)
        if flag:
            return result
        else:
            return error

    def read_attendanceTable(self, attendanceTableFilename):
        wookbookAttendance = openpyxl.load_workbook(attendanceTableFilename)
        worksheetAttendance = wookbookAttendance.active
        attendance = []
        for i in range(1, worksheetAttendance.max_row):
            day = []
            j = 0
            for col in worksheetAttendance.iter_cols(1, worksheetAttendance.max_column):
                j += 1
                if (col[i].value != None):
                    day.append(float(col[i].value))
            attendance.append(day)
        return attendance

    def save_table(self, table, newTableFilename):
        wookbookNewTable = openpyxl.load_workbook(newTableFilename)
        worksheetNewTable = wookbookNewTable.active
        worksheetNewTable["A1"] = "Понедельник"
        worksheetNewTable["B1"] = "Вторник"
        worksheetNewTable["C1"] = "Среда"
        worksheetNewTable["D1"] = "Четверг"
        worksheetNewTable["E1"] = "Пятница"
        worksheetNewTable["F1"] = "Суббота"
        for i in range(len(table)):
            for j in range(len(table[i])):
                if table[i][j] == []:
                    worksheetNewTable.cell(i + 2, j + 1, "X")
                else:
                    worksheetNewTable.cell(i + 2, j + 1, table[i][j])
        wookbookNewTable.save(newTableFilename)