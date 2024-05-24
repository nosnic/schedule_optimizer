import json
import openpyxl


class TableWorker:
    class Lesson:
        def __init__(self, name, information, day, time):
            self.name = name
            self.teacher = information.get("teacher")
            self.session = information.get("session")
            self.profile = information.get("profile")
            self.notable = information.get("notable")
            self.lecture = information.get("lecture")
            self.day = day
            self.time = time

    def readonly_table(self, table_file_name):
        workbook_table = openpyxl.load_workbook(table_file_name, read_only=True)
        worksheet_table = workbook_table.active
        table = [
            [
                col.value if col.value is not None else ""
                for col in row
            ]
            for row in worksheet_table.iter_rows(min_row=2)
        ]
        return table

    def read_table(self, table_file_name, additional_info_file_name):
        with open(additional_info_file_name, 'r', encoding='utf-8') as f:
            data = json.load(f)

        workbook_table = openpyxl.load_workbook(table_file_name)
        worksheet_table = workbook_table.active

        table = []
        for i, row in enumerate(worksheet_table.iter_rows(min_row=2), start=1):
            day = []
            for j, cell in enumerate(row, start=1):
                if cell.value is not None and cell.value in data:
                    lesson_info = data[cell.value]
                    lesson_info["day"] = j
                    lesson_info["time"] = i
                    day.append(lesson_info)
            table.append(day)
        return table

    def create_info_table(self, table, additional_info_file_name, saving_lessons=[]):
        with open(additional_info_file_name, 'r', encoding='utf-8') as f:
            data = json.load(f)

        result = []
        flag = True
        error = ""

        valid_keys = set(data.keys())
        for i, day in enumerate(table):
            day_result = []
            for j, lesson in enumerate(day):
                if lesson:
                    if lesson in valid_keys:
                        if [i, j] in saving_lessons:
                            data[lesson] = data.get("Максимум", {})
                        lesson_info = data[lesson]
                        lesson_info["day"] = j
                        lesson_info["time"] = i
                        day_result.append(lesson_info)
                    else:
                        flag = False
                        error = lesson
            result.append(day_result)

        return result if flag else error

    def read_attendance_table(self, attendance_table_filename):
        workbook_attendance = openpyxl.load_workbook(attendance_table_filename, read_only=True)
        worksheet_attendance = workbook_attendance.active

        attendance = [
            [
                float(cell.value) if cell.value is not None else None
                for cell in row
            ]
            for row in worksheet_attendance.iter_rows(min_row=2)
        ]
        return attendance

    def save_table(self, table, new_table_filename):
        workbook_new_table = openpyxl.Workbook()
        worksheet_new_table = workbook_new_table.active

        days = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]
        for col_num, day in enumerate(days, start=1):
            worksheet_new_table.cell(row=1, column=col_num, value=day)

        for i, day in enumerate(table, start=2):
            for j, lesson in enumerate(day, start=1):
                value = "X" if not lesson else lesson
                worksheet_new_table.cell(row=i, column=j, value=value)

        workbook_new_table.save(new_table_filename)
